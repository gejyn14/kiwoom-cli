#!/usr/bin/env python3
"""Regenerate kiwoom_cli_commands.xlsx from current CLI definitions.

Usage: python scripts/regen_excel.py
"""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

import click
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from kiwoom_cli.main import cli

GROUP_COLORS = {
    "config": "E2EFDA", "auth": "E2EFDA", "api": "E2EFDA",
    "stock": "DAEEF3", "account": "FFF2CC", "order": "FCE4D6",
    "market": "E4DFEC", "stream": "D6E4F0",
    "dashboard": "F2F2F2", "watch": "F2F2F2",
}

GROUP_INFO = {
    "config": "설정 관리", "auth": "인증", "api": "Raw API",
    "stock": "종목 조회", "account": "계좌 조회", "order": "주문",
    "market": "시장 정보", "stream": "실시간 스트리밍",
    "dashboard": "대시보드", "watch": "실시간 모니터링",
}


def collect_all(group, prefix: str = "kiwoom") -> list[dict]:
    """Recursively extract all CLI commands from a Click group."""
    results = []
    ctx = click.Context(group)
    for name in sorted(group.list_commands(ctx)):
        cmd = group.get_command(ctx, name)
        full = f"{prefix} {name}"
        if isinstance(cmd, click.Group):
            results.extend(collect_all(cmd, full))
        elif isinstance(cmd, click.Command):
            if getattr(cmd, "hidden", False):
                continue
            doc = (cmd.help or cmd.callback.__doc__ or "").strip()
            m = re.search(r"\(([a-z]{2}\d{4,5})\)", doc)
            api_id = m.group(1) if m else ""
            desc = doc.split(".")[0].strip() if doc else ""
            parts = full.split(" ")
            if len(parts) == 3:
                grp, sub, command = parts[1], "-", parts[2]
            elif len(parts) == 4:
                grp, sub, command = parts[1], parts[2], parts[3]
            elif len(parts) >= 5:
                grp, sub, command = parts[1], " ".join(parts[2:-1]), parts[-1]
            else:
                grp, sub, command = parts[1] if len(parts) > 1 else "-", "-", name
            args, opts = [], []
            for p in cmd.params:
                if isinstance(p, click.Argument):
                    nargs = " (복수)" if p.nargs == -1 else ""
                    req = "필수" if p.required else "선택"
                    args.append(f"{p.human_readable_name}{nargs} ({req})")
                elif isinstance(p, click.Option) and "--help" not in p.opts:
                    flag = p.opts[0]
                    info = [flag]
                    if isinstance(p.type, click.Choice):
                        info.append(f"[{', '.join(p.type.choices)}]")
                    d = str(p.default) if p.default is not None else None
                    if d and d != "" and d != "False" and d != "0" and "Sentinel" not in d:
                        info.append(f"기본:{d}")
                    if p.required:
                        info.append("*필수")
                    if p.help:
                        info.append(f"({p.help})")
                    opts.append(" ".join(info))
            results.append({
                "group": grp, "subgroup": sub, "command": command,
                "full_cmd": full, "description": desc, "api_id": api_id,
                "arguments": "\n".join(args) if args else "-",
                "options": "\n".join(opts) if opts else "-",
            })
    return results


def build_workbook(commands: list[dict]) -> Workbook:
    """Build the styled Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "전체 명령어"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    wa = Alignment(vertical="top", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="top")
    headers = ["분류", "하위그룹", "명령어", "전체 명령어", "설명", "API ID", "인자", "옵션"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = tb
    widths = {"A": 10, "B": 12, "C": 20, "D": 45, "E": 32, "F": 10, "G": 22, "H": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i, cmd in enumerate(commands, 2):
        rf = PatternFill(
            start_color=GROUP_COLORS.get(cmd["group"], "FFFFFF"),
            end_color=GROUP_COLORS.get(cmd["group"], "FFFFFF"),
            fill_type="solid",
        )
        values = [
            cmd["group"], cmd["subgroup"], cmd["command"], cmd["full_cmd"],
            cmd["description"], cmd["api_id"], cmd["arguments"], cmd["options"],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = tb
            c.fill = rf
            c.alignment = ca if col == 6 else wa
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(commands) + 1}"

    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="kiwoom-cli 명령어 요약").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="분류").font = Font(bold=True)
    ws2.cell(row=3, column=2, value="명령어 수").font = Font(bold=True)
    ws2.cell(row=3, column=3, value="설명").font = Font(bold=True)
    counts = Counter(c["group"] for c in commands)
    row = 4
    total = 0
    for g in ["config", "auth", "api", "stock", "account", "order", "market", "stream", "dashboard", "watch"]:
        cnt = counts.get(g, 0)
        total += cnt
        rf = PatternFill(
            start_color=GROUP_COLORS.get(g, "FFFFFF"),
            end_color=GROUP_COLORS.get(g, "FFFFFF"),
            fill_type="solid",
        )
        for col, val in enumerate([g, cnt, GROUP_INFO.get(g, "")], 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.fill = rf
            c.border = tb
        row += 1
    ws2.cell(row=row, column=1, value="합계").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 55
    return wb


def main() -> None:
    commands = collect_all(cli)
    wb = build_workbook(commands)
    out = PROJECT_ROOT / "kiwoom_cli_commands.xlsx"
    wb.save(out)
    print(f"Excel 재생성 완료: {out} ({len(commands)}개 명령어)")


if __name__ == "__main__":
    main()
